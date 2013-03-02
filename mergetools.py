# -*- coding: utf-8 -*-
"""
This file is part of Datamerger

Datamerger is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

Datamerger is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

Refer to <http://www.gnu.org/licenses/> for a copy of the GNU General Public License.
"""

import os
import sys
import csv

# Import modules required for Excel files
import xlrd	
import xlwt

# For writing Excel 2007 and higher xlsx
from openpyxl import Workbook


def read_csv(path_to_csv):
	try:
		f_csv = open(path_to_csv,"rb")
	except IOError as e:
		print >> sys.stderr, e
		return False
		
	try:
		dialect = csv.Sniffer().sniff(f_csv.readline())
	except:
		print >> sys.stderr, "Failed to sniff file parameters, assuming , as delimiter symbol"
		dialect = csv.get_dialect('excel')
		
	f_csv.seek(0)			
	data = csv.DictReader(f_csv,dialect=dialect,restkey="UNKNOWN",restval="")
	fieldnames = data.fieldnames
	read_data = []
	
	for row in data:
		#if "UNKNOWN" in row and not "UNKNOWN" in fieldnames:
		#	fieldnames.append("UNKNOWN")
		row["source_file"] = os.path.split(path_to_csv)[1]
		read_data.append(row)
		
	return(fieldnames, read_data)
	
	
	
def write_csv(path_to_csv, header, data):
	try:
		out_fp = open(path_to_csv,"w")
	except IOError as e:
		print >> sys.stderr, e
		return False
		
	output = csv.DictWriter(out_fp, fieldnames=header, dialect='excel', lineterminator="\n", quotechar="\"", restval="",escapechar="\\")
	output.writeheader()
	
	errorCount = 0
	for row in data:
		try:
			output.writerow(row)
		except Exception as e:
			print >> sys.stderr, "Warning ("+ row["source_file"] + "): " + str(e)
			errorCount += 1			
	return errorCount
	
def read_xls(path_to_xls):	
	workbook = xlrd.open_workbook(path_to_xls)
	sheet = workbook.sheet_by_index(0)
	headers = sheet.row_values(0)
	data = []
	
	for row in xrange(1,sheet.nrows):
		row_values = sheet.row_values(row)
		row_data = {}
		for col in xrange(0,len(row_values)):
			row_data[headers[col]] = row_values[col]
			row_data["source_file"] = os.path.split(path_to_xls)[1]
		data.append(row_data)
				
	return (headers, data)

def write_xls(path_to_xls, header, data):
	workbook = xlwt.Workbook(encoding = 'utf8')
	worksheet = workbook.add_sheet("merged_data")

	# Write header
	font = xlwt.Font() # Create the Font
	font.name = 'Arial'
	font.bold = True
	style = xlwt.XFStyle() # Create the Style
	style.font = font # Apply the Font to the Style	
	for col in xrange(len(header)):
		worksheet.write(0,col,header[col],style)
		
	# Write data	
	for row in xrange(0,len(data)):					
		for col in xrange(len(header)):
			col_name = header[col]			
			value = data[row][col_name]								
			worksheet.write(row+1,col,value)
	
	workbook.save(path_to_xls)
	return 0
	
	
def write_xlsx(path_to_xlsx, header, data):
	workbook = Workbook()
	worksheet = workbook.worksheets[0]
	
	worksheet.title = "merged_data"
		
	# Write header
	for col in xrange(len(header)):
		cell = worksheet.cell(row=0,column=col)
		cell.value = header[col]
		cell.style.font.bold = True
		
	# Write data	
	for row in xrange(0,len(data)):					
		for col in xrange(len(header)):
			col_name = header[col]			
			value = data[row][col_name]								
			worksheet.cell(row=row+1,column=col).value = value
	
	workbook.save(filename = str(path_to_xlsx))
	return 0

def mergeFolder(folder, destination, ui=None):
	
	if not os.path.isdir(folder):
		print >> sys.stderr, "ERROR: the specified folder " + folder + " is invalid"
		return False
		
	folder_files = os.listdir(folder)	
	valid_files = filter(lambda x: os.path.splitext(x)[1] in  [".csv",".xls",".xlsx"], folder_files)
	
	if not len(valid_files):
		print >> sys.stderr, "ERROR: Folder contains no mergable files"
		return False

	col_names = []	
	total_data = []
	unknown_found = False
	
	counter = 0	
	for datafile in valid_files:
		print "Reading file " + datafile
		filetype = os.path.splitext(datafile)[1]
		
		if filetype == ".csv":
			(header, data) = read_csv(os.path.join(folder,datafile))										
		elif filetype in [".xls",".xlsx"]:
			(header, data) = read_xls(os.path.join(folder,datafile))						
	
		col_names = list(set(col_names) | set(header) )
		total_data.extend(data)
		counter += 1
						
		if not ui is None:
			progress = int(counter/float(len(valid_files)+1)*100)
			ui.progressBar.setValue(progress)	
	
	col_names = ["source_file"] + col_names
	if unknown_found:
		col_names += ["UNKNOWN"]
		
	print "Writing data to file (please be patient, this can take a while...)"	
	
	destination_ext = os.path.splitext(str(destination))[1]
	
	if destination_ext == ".csv":	
		errorCount = write_csv(destination, col_names, total_data)
	elif destination_ext == ".xls":
		errorCount = write_xls(destination, col_names, total_data)
	elif destination_ext == ".xlsx":
		errorCount = write_xlsx(destination, col_names, total_data)
			
	ui.progressBar.setValue(100)			
	print "Merged " + str(counter) + " files with " + str(errorCount) + " errors."
	return True
	
	