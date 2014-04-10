# -*- coding: utf-8 -*-
"""
This file is part of Datamerger

Datamerger is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

Datamerger is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

Refer to <http://www.gnu.org/licenses/> for a copy of the GNU General Public License.

@author Daniel Schreij
"""

import os
import sys
import csv

# Import modules required for Excel files
import xlrd	
import xlwt

# For writing Excel 2007 and higher xlsx. Older versions of openpyxl may not
# have the Workbook.
try:
	from openpyxl import Workbook
except:
	Workbook = None

def correct_datatype(value):
	""" Convert values to correct datatype if they are int or float
	If they are uncovertible, just return the original string. """
	
	try:
		return int(value)
	except ValueError:
		pass
	try:
		return float(value)
	except ValueError:
		pass
	return value
	

def read_csv(path_to_csv):
	"""
	Reads csv file to a list containing dictionaries, each representing a row. 
	The keys of the dictionary represents the column names, and the value contains
	the corresponding value of that cell.
	
	Args:
		path_to_csv (string): a path to the csv file to be parsed
	Returns:
		a tuple containing a list with column names, and the rest of the data
		as a list of dictionaries

	"""
	try:
		f_csv = open(path_to_csv,"rb")
	except IOError as e:
		print >> sys.stderr, e
		return False
		
	try:
		dialect = csv.Sniffer().sniff(f_csv.readline())
	except:
		print >> sys.stderr, "Failed to sniff file parameters, assuming , as delimiter symbol"
		dialect = csv.get_dialect('excel')
		
	f_csv.seek(0)			
	data = csv.DictReader(f_csv,dialect=dialect,restkey="UNKNOWN",restval="")
	fieldnames = data.fieldnames
	read_data = []
	
	for row in data:
		row["dm_source_file"] = os.path.split(path_to_csv)[1]
		read_data.append(row)
		
	return(fieldnames, read_data)
	
	
def write_csv(path_to_csv, header, data, ui=None, files=None):
	"""
	Writes a csv file from the specified data
	
	Args:
		path_to_csv (string): The path where to save the csv file
		header (list): A list with the header names (as strings)
		data (list): A list with dictionaries which each represent a row of the data
			The dictionary keys should represent the column names and the corresponding values
			represent the values for the current row
	Returns:
		errorCount (int): 0 if no errors, or otherwise the number of (non-critical) errors that occurred
	"""
	try:
		out_fp = open(path_to_csv,"w")
	except IOError as e:
		print >> sys.stderr, e
		return False
		
	# Sort header alphabetically	
	header.sort()	
	
	output = csv.DictWriter(out_fp, fieldnames=header, dialect='excel', lineterminator="\n", quotechar="\"", restval="",escapechar="\\")
	output.writeheader()
	
	errorCount = 0
	counter = 0
	if ui and files:
		rows_part = len(data)/files	
	
	for row in data:
		counter += 1
		try:
			output.writerow(row)
		except Exception as e:
			print >> sys.stderr, "Warning ("+ row["dm_source_file"] + "): " + str(e)
			errorCount += 1
		
		if not ui is None:
			if counter % rows_part == 0:
				part = counter/rows_part								
				progress = int((part+files)/float(2*files+1)*100)
				ui.progressBar.setValue(progress)	
				
			
	return errorCount
	
def read_xls(path_to_xls):
	"""
	Reads .xls or .xlsx file to a list containing dictionaries, each representing a row. 
	The keys of the dictionary represents the column names, and the value contains
	the corresponding value of that cell.
	
	Args:
		path_to_csv (string): a path to the csv file to be parsed
	Returns:
		a tuple containing a list with column names, and the rest of the data
		represented as a list of dictionaries

	"""
	workbook = xlrd.open_workbook(path_to_xls)
	sheet = workbook.sheet_by_index(0)
	headers = sheet.row_values(0)
	data = []
	
	for row in xrange(1,sheet.nrows):
		row_values = sheet.row_values(row)
		row_data = {}
		for col in xrange(0,len(row_values)):
			row_data[headers[col]] = row_values[col]
			row_data["dm_source_file"] = os.path.split(path_to_xls)[1]
		data.append(row_data)
				
	return (headers, data)

def write_xls(path_to_xls, header, data, ui=None, files=None):
	"""
	Writes a xls (Excel 97-2003) file from the specified data
	
	Args:
		path_to_csv (string): The path where to save the csv file
		header (list): A list with the header names (as strings)
		data (list): A list with dictionaries which each represent a row of the data
			The dictionary keys should represent the column names and the corresponding values
			represent the values for the current row
	Returns:
		errorCount (int): 0 if no errors, or otherwise the number of (non-critical) errors that occurred
	"""
	# Sort header alphabetically	
	header.sort()	
	
	workbook = xlwt.Workbook(encoding = 'utf8')
	worksheet = workbook.add_sheet("merged_data")

	# Write header in bold
	font = xlwt.Font() # Create the Font
	font.name = 'Arial'
	font.bold = True
	style = xlwt.XFStyle() # Create the Style
	style.font = font # Apply the Font to the Style	
	for col in xrange(len(header)):
		worksheet.write(0,col,header[col],style)

	if ui and files:
		rows_part = len(data)/files			
		
	# Write data	
	for row in xrange(0,len(data)):				
		for col in xrange(len(header)):
			col_name = header[col]	
			try:
				value = data[row][col_name]
			except KeyError:
				value = ""		
						
			worksheet.write(row+1, col, correct_datatype(value) )
		
		if not ui is None:
			if row % rows_part == 0:
				part = row/rows_part								
				progress = int((part+files)/float(2*files+1)*100)
				ui.progressBar.setValue(progress)	
	
	workbook.save(path_to_xls)
	return 0
	
	
def write_xlsx(path_to_xlsx, header, data, ui=None, files=None):
	"""
	Writes a xlsx file (Excel 2010 and higher) from the specified data
	
	Args:
		path_to_csv (string): The path where to save the csv file
		header (list): A list with the header names (as strings)
		data (list): A list with dictionaries which each represent a row of the data
			The dictionary keys should represent the column names and the corresponding values
			represent the values for the current row
	Returns:
		errorCount (int): 0 if no errors, or otherwise the number of (non-critical) errors that occurred
	"""

	# Check if Workbook functionality is available. TODO handle this in a more
	# elegant way.
	if Workbook is None:
		return 1
		
	# Sort header alphabetically	
	header.sort()	
	
	workbook = Workbook(optimized_write = True)
	worksheet = workbook.create_sheet()
	
	worksheet.title = "merged_data"
				
	if ui and files:
		rows_part = len(data)/files	
	
	### Write data	

	# Write column names first
	worksheet.append(header)	
	
	for row in xrange(0,len(data)):	
		row_data = []				
		for col in xrange(len(header)):
			col_name = header[col]	
			try:
				value = data[row][col_name]
			except KeyError:
				value = ""
			row_data.append(correct_datatype(value))
			
		worksheet.append(row_data)
		
		if not ui is None:
			if row % rows_part == 0:
				part = row/rows_part								
				progress = int((part+files)/float(2*files+1)*100)
				ui.progressBar.setValue(progress)	
	
	workbook.save(filename = str(path_to_xlsx))
	return 0

def mergeFolder(folder, destination, ui=None):
	"""
	Merges the spreadsheet files in the specified folder to one spreadsheet in one file
	
	Args:
		folder (string): The folder to scan for spreadsheet (csv,xls,xlsx) files
		destination (string): The path and filename to save the merged data to (extension should be csv,xls,xlsx)
		ui (PyQt.QtGui.QWidget) [optional]: If relevant, a reference to the Datamerger UI.
	Returns:
		errorCount (int): 0 if no errors, or otherwise the number of (non-critical) errors that occurred
	"""	
	# Check if folder points to an existing dir
	if not os.path.isdir(folder):
		print >> sys.stderr, "ERROR: the specified folder " + folder + " is invalid"
		return False
		
	# Filter for files with compatible extensions
	folder_files = os.listdir(folder)	
	valid_files = filter(lambda x: os.path.splitext(x)[1] in  [".csv",".xls",".xlsx"], folder_files)
	
	if not len(valid_files):
		print >> sys.stderr, "ERROR: Folder contains no mergable files"
		return False

	col_names = []	
	total_data = []
	unknown_found = False
	
	counter = 0	
	for datafile in valid_files:
		print "Reading file " + datafile
		filetype = os.path.splitext(datafile)[1]
		
		if filetype == ".csv":
			(header, data) = read_csv(os.path.join(folder,datafile))										
		elif filetype in [".xls",".xlsx"]:
			(header, data) = read_xls(os.path.join(folder,datafile))						
	
		col_names = list(set(col_names) | set(header) )
		total_data.extend(data)
		counter += 1
						
		if not ui is None:
			progress = int(counter/float(2*len(valid_files)+1)*100)
			ui.progressBar.setValue(progress)	
	
	print "Writing merged data to file (please be patient as this can take a while...)"
	#ui.progressBar.setValue(progress+1)  #If this is ommitted, above line is not printed to textbox in GUI...
	
	# Add column to add source file info
	col_names = ["dm_source_file"] + col_names
	
	# Add column for unnamed columns (only applicable for incorrectly formatted csv's.
	# Usually these files crash the process, as they should)
	if unknown_found:
		col_names += ["UNKNOWN"]
	
	destination_ext = os.path.splitext(str(destination))[1]
	
	if destination_ext == ".csv":	
		errorCount = write_csv(destination, col_names, total_data, ui, len(valid_files))
	elif destination_ext == ".xls":
		errorCount = write_xls(destination, col_names, total_data, ui, len(valid_files))
	elif destination_ext == ".xlsx":
		errorCount = write_xlsx(destination, col_names, total_data, ui, len(valid_files))
	
	if not ui is None:
		ui.progressBar.setValue(100)			

	print "Merged " + str(counter) + " files with " + str(errorCount) + " errors."
	
	return True
	
	